import re
import io
from datetime import datetime
import traceback
import pandas as pd
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt
from app.extensions import db
from app.models import Student, Company, User

students_bp = Blueprint('students', __name__)

def parse_float_safe(val, default=None):
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ('none', 'null', 'nan', 'n/a', ''):
        return default
    try:
        return float(val_str)
    except (ValueError, TypeError):
        return default

def parse_int_safe(val, default=None):
    if val is None:
        return default
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ('none', 'null', 'nan', 'n/a', ''):
        return default
    try:
        return int(float(val_str))
    except (ValueError, TypeError):
        return default

@students_bp.route('/', methods=['GET'])
def get_students():
    try:
        department = request.args.get('department')
        gender = request.args.get('gender')
        hosteller = request.args.get('hosteller_status')
        placement_status = request.args.get('placement_status')
        search = request.args.get('search')

        query = Student.query
        if department:
            query = query.filter(Student.department.ilike(f'%{department}%'))
        if gender:
            query = query.filter(Student.gender == gender)
        if hosteller:
            query = query.filter(Student.hosteller_status == hosteller)
        if placement_status:
            if placement_status.upper() == 'PLACED':
                query = query.filter(Student.placement_status == 'PLACED')
            else:
                query = query.filter((Student.placement_status != 'PLACED') | (Student.placement_status.is_(None)))
        if search:
            query = query.filter(
                (Student.name.ilike(f'%{search}%')) |
                (Student.reg_no.ilike(f'%{search}%'))
            )

        students = query.order_by(Student.s_no.asc(), Student.id.asc()).all()
        return jsonify([s.to_summary_dict() for s in students]), 200
    except Exception as e:
        print(f"[ERROR] get_students failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to retrieve students', 'details': str(e)}), 500

@students_bp.route('/<int:student_id>', methods=['GET'])
def get_student_details(student_id):
    try:
        student = db.session.get(Student, student_id)
        if not student:
            return jsonify({'error': 'Student not found'}), 404
        return jsonify(student.to_full_dict()), 200
    except Exception as e:
        print(f"[ERROR] get_student_details failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to retrieve student details', 'details': str(e)}), 500

@students_bp.route('/', methods=['POST'])
def add_student():
    try:
        data = request.get_json() or {}
        
        # 1. Reg No & Name Validation
        reg_no = str(data.get('reg_no') or '').strip()
        name = str(data.get('name') or '').strip()
        
        if not reg_no:
            return jsonify({'error': 'Registration number is required', 'details': 'reg_no field cannot be empty'}), 400
        if not name:
            return jsonify({'error': 'Student name is required', 'details': 'name field cannot be empty'}), 400

        # Check duplicate
        existing = Student.query.filter(Student.reg_no.ilike(reg_no)).first()
        if existing:
            return jsonify({
                'error': 'Student with this registration number already exists.',
                'details': f'Registration number {reg_no} is already registered to {existing.name}.'
            }), 409

        # 2. Percentage validations (0 to 100)
        sslc_pct = parse_float_safe(data.get('sslc_percentage') if data.get('sslc_percentage') is not None else data.get('tenth_percentage'), default=0.0)
        hsc_pct = parse_float_safe(data.get('hsc_percentage') if data.get('hsc_percentage') is not None else data.get('twelfth_percentage'), default=0.0)
        ug_pct = parse_float_safe(data.get('ug_percentage') if data.get('ug_percentage') is not None else data.get('cgpa'), default=0.0)
        pg_pct = parse_float_safe(data.get('pg_percentage'), default=None)
        diploma_pct = parse_float_safe(data.get('diploma_percentage'), default=0.0)

        for pct_name, pct_val in [('SSLC %', sslc_pct), ('HSC %', hsc_pct), ('UG %', ug_pct), ('Diploma %', diploma_pct)]:
            if pct_val is not None and (pct_val < 0 or pct_val > 100):
                return jsonify({'error': f'{pct_name} must be between 0 and 100'}), 400

        if pg_pct is not None and (pg_pct < 0 or pg_pct > 100):
            return jsonify({'error': 'PG % must be between 0 and 100'}), 400

        # 3. Arrears validations (non-negative)
        curr_arr = parse_int_safe(data.get('current_arrears'), default=0)
        hist_arr = parse_int_safe(data.get('history_arrears'), default=0)
        if curr_arr < 0:
            return jsonify({'error': 'Current arrears cannot be negative'}), 400
        if hist_arr < 0:
            return jsonify({'error': 'History of arrears cannot be negative'}), 400

        # 4. Links validation (safe format check)
        for link_field in ['github_id', 'linkedin_id', 'resume_link', 'self_intro_link', 'photo_link', 'portfolio_link']:
            link_val = str(data.get(link_field) or '').strip()
            if link_val and not (link_val.startswith('http://') or link_val.startswith('https://') or '.' in link_val):
                return jsonify({'error': f'Invalid URL format for {link_field}'}), 400

        # 5. Email & Phone format checks if present
        email = str(data.get('email') or '').strip()
        if email:
            email_pattern = r'^[^@\s]+@[^@\s]+\.[^@\s]+$'
            if not re.match(email_pattern, email):
                return jsonify({'error': 'Invalid email format'}), 400

        phone = str(data.get('phone') or data.get('mobile_no') or '').strip()
        if phone:
            digits_only = re.sub(r'\D', '', phone)
            if len(digits_only) < 7 or len(digits_only) > 15:
                return jsonify({'error': 'Phone number must be between 7 and 15 digits'}), 400

        # Auto-compute next s_no
        max_sno = db.session.query(db.func.max(Student.s_no)).scalar() or 0
        new_sno = max_sno + 1

        student = Student(
            s_no=new_sno,
            reg_no=reg_no,
            name=name,
            department=str(data.get('department') or data.get('dept') or 'CSE').strip(),
            gender=str(data.get('gender') or 'Male').strip(),
            hosteller_status=str(data.get('hosteller_status') or data.get('hosteller_day_scholar') or 'Day Scholar').strip(),
            sslc_percentage=sslc_pct,
            hsc_percentage=hsc_pct,
            ug_percentage=ug_pct,
            pg_percentage=pg_pct,
            diploma_percentage=diploma_pct,
            current_arrears=curr_arr,
            history_arrears=hist_arr,
            graduation_year=parse_int_safe(data.get('graduation_year'), default=None),
            github_id=str(data.get('github_id') or '').strip(),
            linkedin_id=str(data.get('linkedin_id') or '').strip(),
            resume_link=str(data.get('resume_link') or '').strip(),
            self_intro_link=str(data.get('self_intro_link') or '').strip(),
            photo_link=str(data.get('photo_link') or '').strip(),
            portfolio_link=str(data.get('portfolio_link') or '').strip(),
            email=email,
            phone=phone,
            placement_status=str(data.get('placement_status') or 'NOT PLACED').strip(),
            placed_company=str(data.get('placed_company') or 'N/A').strip(),
            salary_package=str(data.get('salary_package') or 'N/A').strip(),
            remarks=str(data.get('remarks') or '').strip()
        )

        db.session.add(student)
        db.session.flush()
        student_dict = student.to_full_dict()
        db.session.commit()
        return jsonify({'message': 'Student added successfully', 'student': student_dict}), 201

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] add_student failed: {traceback.format_exc()}")
        return jsonify({'error': 'Unable to save student details. Please verify your connection and try again.', 'details': str(e)}), 500

@students_bp.route('/<int:student_id>', methods=['PUT'])
def update_student(student_id):
    try:
        student = db.session.get(Student, student_id)
        if not student:
            return jsonify({'error': 'Student not found'}), 404

        data = request.get_json() or {}

        if 'name' in data:
            student.name = str(data['name']).strip()
        if 'department' in data or 'dept' in data:
            student.department = str(data.get('department') or data.get('dept')).strip()
        if 'gender' in data:
            student.gender = str(data['gender']).strip()
        if 'hosteller_status' in data or 'hosteller_day_scholar' in data:
            student.hosteller_status = str(data.get('hosteller_status') or data.get('hosteller_day_scholar')).strip()
        if 'sslc_percentage' in data or 'tenth_percentage' in data:
            student.sslc_percentage = parse_float_safe(data.get('sslc_percentage') if 'sslc_percentage' in data else data.get('tenth_percentage'), default=0.0)
        if 'hsc_percentage' in data or 'twelfth_percentage' in data:
            student.hsc_percentage = parse_float_safe(data.get('hsc_percentage') if 'hsc_percentage' in data else data.get('twelfth_percentage'), default=0.0)
        if 'ug_percentage' in data or 'cgpa' in data:
            student.ug_percentage = parse_float_safe(data.get('ug_percentage') if 'ug_percentage' in data else data.get('cgpa'), default=0.0)
        if 'pg_percentage' in data:
            student.pg_percentage = parse_float_safe(data['pg_percentage'], default=None)
        if 'diploma_percentage' in data:
            student.diploma_percentage = parse_float_safe(data['diploma_percentage'], default=0.0)
        if 'current_arrears' in data:
            student.current_arrears = parse_int_safe(data['current_arrears'], default=0)
        if 'history_arrears' in data:
            student.history_arrears = parse_int_safe(data['history_arrears'], default=0)
        if 'graduation_year' in data:
            student.graduation_year = parse_int_safe(data['graduation_year'], default=None)
        if 'github_id' in data:
            student.github_id = str(data['github_id']).strip()
        if 'linkedin_id' in data:
            student.linkedin_id = str(data['linkedin_id']).strip()
        if 'resume_link' in data:
            student.resume_link = str(data['resume_link']).strip()
        if 'self_intro_link' in data:
            student.self_intro_link = str(data['self_intro_link']).strip()
        if 'photo_link' in data:
            student.photo_link = str(data['photo_link']).strip()
        if 'portfolio_link' in data:
            student.portfolio_link = str(data['portfolio_link']).strip()
        if 'email' in data:
            student.email = str(data['email']).strip()
        if 'phone' in data or 'mobile_no' in data:
            student.phone = str(data.get('phone') or data.get('mobile_no')).strip()
        if 'placement_status' in data:
            student.placement_status = str(data['placement_status']).strip()
        if 'placed_company' in data:
            student.placed_company = str(data['placed_company']).strip()
        if 'salary_package' in data:
            student.salary_package = str(data['salary_package']).strip()
        if 'remarks' in data:
            student.remarks = str(data['remarks']).strip()

        db.session.commit()
        return jsonify({'message': 'Student updated successfully', 'student': student.to_full_dict()}), 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] update_student failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to update student', 'details': str(e)}), 500

def get_current_user_role():
    try:
        claims = get_jwt()
        if claims and 'role' in claims:
            return claims['role']
    except Exception:
        pass
    try:
        user_id = get_jwt_identity()
        if user_id:
            user = db.session.get(User, int(user_id))
            if user:
                return user.role
    except Exception:
        pass
    return None

def resequence_student_sno():
    """Ensure all student records in database have clean, continuous 1..N s_no with no gaps"""
    try:
        all_students = Student.query.order_by(Student.s_no.asc().nulls_last(), Student.id.asc()).all()
        for idx, s in enumerate(all_students, start=1):
            if s.s_no != idx:
                s.s_no = idx
        db.session.flush()
    except Exception as e:
        print(f"[WARN] Failed to resequence student s_no: {e}")

@students_bp.route('/<int:student_id>', methods=['DELETE'])
@jwt_required()
def delete_student(student_id):
    role = get_current_user_role()
    if role != 'ADMIN':
        return jsonify({'error': 'Forbidden: Only Admin has authority to delete students.'}), 403

    try:
        student = db.session.get(Student, student_id)
        if not student:
            return jsonify({'error': 'Student not found.'}), 404

        company_id = student.placed_company_id
        db.session.delete(student)
        db.session.flush()

        # If student was placed, synchronize the company's real placed count
        if company_id:
            comp = db.session.get(Company, company_id)
            if comp:
                comp.placed_students = comp.get_real_placed_count()

        # Automatically re-sequence s_no in order for all remaining students
        resequence_student_sno()

        db.session.commit()
        return jsonify({'message': f'Student "{student.name}" deleted successfully.'}), 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] delete_student failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to delete student.', 'details': str(e)}), 500

@students_bp.route('/bulk', methods=['DELETE'])
@jwt_required()
def bulk_delete_students():
    """Admin exclusive endpoint to bulk delete multiple students"""
    role = get_current_user_role()
    if role != 'ADMIN':
        return jsonify({'error': 'Forbidden: Only Admin has authority to bulk delete students.'}), 403

    try:
        data = request.get_json() or {}
        student_ids = data.get('student_ids', [])

        if not student_ids or not isinstance(student_ids, list):
            return jsonify({'error': 'Please provide a valid list of student IDs to delete.'}), 400

        # Find matching students
        students = Student.query.filter(Student.id.in_(student_ids)).all()
        if not students:
            return jsonify({'error': 'No matching students found to delete.'}), 404

        # Track affected companies for live count recalculation
        affected_company_ids = set()
        for student in students:
            if student.placed_company_id:
                affected_company_ids.add(student.placed_company_id)
            db.session.delete(student)

        db.session.flush()

        # Synchronize placement counts for all affected companies
        for cid in affected_company_ids:
            comp = db.session.get(Company, cid)
            if comp:
                comp.placed_students = comp.get_real_placed_count()

        # Automatically re-sequence s_no in order for all remaining students
        resequence_student_sno()

        db.session.commit()
        deleted_count = len(students)
        return jsonify({
            'message': f'{deleted_count} students deleted successfully.',
            'deleted_count': deleted_count
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] bulk_delete_students failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to bulk delete selected students.', 'details': str(e)}), 500

@students_bp.route('/<int:student_id>/placement', methods=['POST'])
@jwt_required()
def place_student(student_id):
    """Mark a student as PLACED in an approved 'Drive Completed' company"""
    try:
        student = db.session.get(Student, student_id)
        if not student:
            return jsonify({'error': 'Student not found.'}), 404

        # Check duplicate placement
        if student.get_norm_placement_status() == 'PLACED':
            return jsonify({'error': 'Student is already placed.'}), 400

        data = request.get_json() or {}
        company_id = data.get('company_id')
        if not company_id:
            return jsonify({'error': 'Company selection is required.'}), 400

        try:
            company_id = int(company_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid company ID.'}), 400

        company = db.session.get(Company, company_id)
        if not company:
            return jsonify({'error': 'Selected company not found.'}), 404

        # Verification: Company must be APPROVED and Drive Completed
        approval_st = (company.approval_status or 'PENDING').upper()
        rel_st = company.status or 'Cold'

        if approval_st != 'APPROVED' or rel_st != 'Drive Completed':
            return jsonify({'error': 'This company is not an approved completed placement drive.'}), 400

        # Hiring limit check
        placed_count = Student.query.filter(
            Student.placed_company_id == company.id,
            Student.placement_status == 'PLACED'
        ).count()

        hirings_limit = company.no_of_hirings if company.no_of_hirings is not None else (company.employee_count or 0)
        if hirings_limit > 0 and placed_count >= hirings_limit:
            return jsonify({'error': "This company's hiring limit has been reached."}), 400

        # Update student placement
        ctc_val = company.ctc_lpa if company.ctc_lpa is not None else 0.0
        placement_date_str = datetime.utcnow().strftime('%d-%m-%Y')

        student.placement_status = 'PLACED'
        student.placed_company_id = company.id
        student.placed_company = company.name
        student.placed_ctc_lpa = ctc_val
        student.salary_package = f"{ctc_val} LPA"
        student.placement_date = placement_date_str

        # Synchronize company's placed_students count
        company.placed_students = placed_count + 1

        db.session.commit()

        return jsonify({
            'message': f"Student successfully placed at {company.name}.",
            'student': student.to_full_dict(),
            'company': company.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] place_student failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to complete student placement.', 'details': str(e)}), 500

@students_bp.route('/upload', methods=['POST'])
def upload_excel():
    """
    Import students from first sheet of Excel file.
    Supports the 100_Students_List.xlsx column structure.
    Second sheet (Placements & Drives) is ALWAYS ignored.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded', 'details': 'file field is missing in request'}), 400

    file = request.files['file']
    filename = file.filename.lower()

    if not (filename.endswith('.xlsx') or filename.endswith('.xls') or filename.endswith('.csv')):
        return jsonify({'error': 'Invalid file type. Please upload an Excel (.xlsx, .xls) or CSV file.'}), 400

    # --- CANONICAL column header mappings (Excel header → student field) ---
    COLUMN_ALIASES = {
        # Roll No / Reg No
        'reg_no':              ['Roll No', 'Roll Number', 'Reg No', 'Register No', 'Registration Number', 'reg_no', 'Reg. No', 'Reg.No'],
        'name':                ['Name', 'Student Name', 'Full Name', 'name'],
        'department':          ['Department', 'Dept', 'dept', 'department'],
        'gender':              ['Gender', 'gender'],
        'hosteller_status':    ['Student Type', 'Hosteller/Day Scholar', 'Hosteller / Day Scholar', 'hosteller/day_scholar',
                                'Residence', 'Hosteller Status', 'hosteller_status', 'Type'],
        'sslc_percentage':     ['SSLC %', '10th %', 'SSLC Percentage', '10th Percentage', 'sslc_percentage', 'tenth_percentage', 'SSLC%'],
        'hsc_percentage':      ['HSC %', '12th %', 'HSC Percentage', '12th Percentage', 'hsc_percentage', 'twelfth_percentage', 'HSC%'],
        'ug_percentage':       ['UG %', 'CGPA', 'UG Percentage', 'ug_percentage', 'cgpa', 'UG%'],
        'pg_percentage':       ['PG %', 'PG Percentage', 'pg_percentage', 'PG%'],
        'diploma_percentage':  ['Diploma %', 'Diploma Percentage', 'diploma_percentage', 'Diploma%'],
        'current_arrears':     ['Current Arrears', 'Arrears', 'current_arrears'],
        'history_arrears':     ['History of Arrears', 'History Arrears', 'history_arrears'],
        'graduation_year':     ['Graduation Date', 'Graduation Year', 'Batch', 'Year of Passing', 'graduation_year', 'Graduation'],
        'github_id':           ['GitHub ID', 'GitHub', 'github_id', 'Github ID'],
        'linkedin_id':         ['LinkedIn ID', 'LinkedIn', 'linkedin_id', 'Linkedin ID'],
        'resume_link':         ['Resume Link', 'Resume', 'resume_link'],
        'portfolio_link':      ['Portfolio', 'Portfolio Link', 'portfolio_link'],
        'email':               ['Personal Email ID', 'Email', 'Student Email', 'email', 'Personal Email', 'Email ID'],
        'college_email':       ['College Email ID', 'College Email', 'college_email', 'Institutional Email'],
        'phone':               ['Mobile No', 'Phone', 'Mobile', 'phone', 'Mobile Number', 'Contact No'],
        'photo_link':          ['Student Photo', 'Photo', 'Photo Link', 'photo_link', 'Student Photo Link'],
        'placement_status':    ['Placement Status', 'placement_status', 'Status'],
        'self_intro_link':     ['Self Intro Link', 'Self Intro Video', 'self_intro_link'],
    }

    def normalize_col(col):
        """Strip and lower-case a column header for fuzzy matching"""
        return str(col).strip().lower()

    def find_col(headers_lower, aliases):
        """Find the first matching header index from a list of aliases"""
        for alias in aliases:
            a_lower = normalize_col(alias)
            if a_lower in headers_lower:
                return headers_lower[a_lower]
        return None

    def cell_val(row, col_idx, default=None):
        """Safe row value getter by positional index"""
        if col_idx is None:
            return default
        try:
            val = row.iloc[col_idx]
            if pd.isna(val):
                return default
            return val
        except Exception:
            return default

    def str_val(row, col_idx, default=''):
        v = cell_val(row, col_idx)
        if v is None:
            return default
        return str(v).strip()

    def normalize_hosteller(raw):
        """Hostel → Hosteller, Day Scholar → Day Scholar"""
        v = str(raw or '').strip()
        v_up = v.upper()
        if v_up in ('HOSTEL', 'HOSTELLER', 'HOSTELITE', 'HOSTELITE'):
            return 'Hosteller'
        if v_up in ('DAY SCHOLAR', 'DAY-SCHOLAR', 'DAYSCHOLAR', 'NON-HOSTELLER'):
            return 'Day Scholar'
        # Partial match
        if 'HOSTEL' in v_up:
            return 'Hosteller'
        if 'DAY' in v_up:
            return 'Day Scholar'
        return v if v else 'Day Scholar'

    def normalize_placement_status(raw):
        """
        Safe placement status normalization.
        Only return 'PLACED' if the value explicitly says PLACED.
        Everything else → 'YET_TO_BE_PLACED'.
        """
        v = str(raw or '').strip().upper()
        if v in ('PLACED', 'YES', 'PLACED - YES'):
            return 'PLACED'
        return 'YET_TO_BE_PLACED'

    def extract_graduation_year(raw):
        """Extract a 4-digit year from 'Graduation Date' which may be a date string or year int"""
        if raw is None:
            return None
        if isinstance(raw, int):
            return raw if 1990 <= raw <= 2050 else None
        if isinstance(raw, float):
            yr = int(raw)
            return yr if 1990 <= yr <= 2050 else None
        # Try parsing date strings like '2024', '2024-05-01', 'May 2024', etc.
        s = str(raw).strip()
        # Try direct 4-digit year
        import re as _re
        m = _re.search(r'\b(19|20)\d{2}\b', s)
        if m:
            return int(m.group(0))
        # Try pandas date parsing
        try:
            import pandas as _pd
            dt = _pd.to_datetime(s, dayfirst=True, errors='coerce')
            if not _pd.isna(dt):
                return dt.year
        except Exception:
            pass
        return None

    try:
        file_bytes = file.read()
        import io as _io

        if filename.endswith('.csv'):
            df_raw = pd.read_csv(_io.BytesIO(file_bytes), header=None, dtype=str)
            sheet_name = 'CSV'
        else:
            # Read ONLY the first sheet; ignore all other sheets (e.g., "Placements & Drives")
            xl = pd.ExcelFile(_io.BytesIO(file_bytes))
            first_sheet = xl.sheet_names[0]
            df_raw = xl.parse(first_sheet, header=None, dtype=str)

        # --- FIND THE ACTUAL HEADER ROW ---
        # Scan up to first 10 rows to locate the row that contains 'Roll No' or 'Reg No' or 'Name'
        HEADER_CLUES = {'roll no', 'reg no', 'register no', 'name', 'student name', 'roll number'}
        header_row_idx = None
        for i in range(min(10, len(df_raw))):
            row_vals = [normalize_col(c) for c in df_raw.iloc[i].fillna('').tolist()]
            if any(clue in row_vals for clue in HEADER_CLUES):
                header_row_idx = i
                break

        if header_row_idx is None:
            return jsonify({
                'error': 'Could not detect header row in the Excel file.',
                'details': 'The importer expects a header row containing "Roll No", "Name", or similar columns within the first 10 rows.'
            }), 400

        # Extract headers and data rows
        headers_raw = df_raw.iloc[header_row_idx].fillna('').tolist()
        headers_lower = {normalize_col(h): idx for idx, h in enumerate(headers_raw)}
        df_data = df_raw.iloc[header_row_idx + 1:].reset_index(drop=True)

        # --- MAP each field to its column index ---
        col_map = {}
        for field, aliases in COLUMN_ALIASES.items():
            col_map[field] = find_col(headers_lower, aliases)

        # reg_no and name are mandatory
        if col_map['reg_no'] is None or col_map['name'] is None:
            return jsonify({
                'error': 'Required columns not found.',
                'details': f'Could not locate "Roll No" or "Name" columns. Found headers: {list(headers_raw)}'
            }), 400

        added_count = 0
        updated_count = 0
        skipped_count = 0
        row_errors = []
        max_sno = db.session.query(db.func.max(Student.s_no)).scalar() or 0

        for row_idx, row in df_data.iterrows():
            # Skip completely blank rows
            if row.dropna().empty or all(str(v).strip() == '' for v in row.tolist()):
                continue

            reg_no = str_val(row, col_map['reg_no']).strip()
            name   = str_val(row, col_map['name']).strip()

            # Skip if both reg_no and name are empty
            if not reg_no or not name:
                skipped_count += 1
                continue

            # Clean up reg_no: remove trailing .0 from float conversion
            if reg_no.endswith('.0'):
                reg_no = reg_no[:-2]

            # --- Build field values ---
            dept = str_val(row, col_map['department'], 'CSE') or 'CSE'
            gender = str_val(row, col_map['gender'], 'Male') or 'Male'
            hosteller_raw = str_val(row, col_map['hosteller_status'], 'Day Scholar')
            hosteller = normalize_hosteller(hosteller_raw)

            sslc_pct  = parse_float_safe(str_val(row, col_map['sslc_percentage']), 0.0)
            hsc_pct   = parse_float_safe(str_val(row, col_map['hsc_percentage']), 0.0)
            ug_pct    = parse_float_safe(str_val(row, col_map['ug_percentage']), 0.0)
            pg_pct    = parse_float_safe(str_val(row, col_map['pg_percentage']), None)
            diploma_pct = parse_float_safe(str_val(row, col_map['diploma_percentage']), 0.0)
            curr_arr  = parse_int_safe(str_val(row, col_map['current_arrears']), 0)
            hist_arr  = parse_int_safe(str_val(row, col_map['history_arrears']), 0)
            grad_year = extract_graduation_year(cell_val(row, col_map['graduation_year']))

            github_id     = str_val(row, col_map['github_id'])
            linkedin_id   = str_val(row, col_map['linkedin_id'])
            resume_link   = str_val(row, col_map['resume_link'])
            portfolio_link = str_val(row, col_map['portfolio_link'])
            photo_link    = str_val(row, col_map['photo_link'])
            # self_intro_link is NOT in the Excel template → always NULL/empty
            self_intro_link = ''

            # Email: prefer Personal Email ID column; fall back to College Email ID
            email = str_val(row, col_map['email'])
            if not email:
                email = str_val(row, col_map['college_email'])

            phone = str_val(row, col_map['phone'])
            # Placement status: safe normalization
            placement_status_raw = str_val(row, col_map['placement_status'], 'YET_TO_BE_PLACED')
            placement_status = normalize_placement_status(placement_status_raw)

            # Check duplicate by reg_no
            existing = Student.query.filter(Student.reg_no.ilike(reg_no)).first()

            try:
                if existing:
                    # Update existing student (do NOT overwrite placement records if already PLACED via workflow)
                    existing.name = name
                    existing.department = dept
                    existing.gender = gender
                    existing.hosteller_status = hosteller
                    existing.sslc_percentage = sslc_pct
                    existing.hsc_percentage = hsc_pct
                    existing.ug_percentage = ug_pct
                    existing.pg_percentage = pg_pct
                    existing.diploma_percentage = diploma_pct
                    existing.current_arrears = curr_arr
                    existing.history_arrears = hist_arr
                    existing.graduation_year = grad_year
                    existing.github_id = github_id
                    existing.linkedin_id = linkedin_id
                    existing.resume_link = resume_link
                    existing.self_intro_link = self_intro_link
                    existing.photo_link = photo_link
                    existing.portfolio_link = portfolio_link
                    existing.email = email
                    existing.phone = phone
                    # Only overwrite placement status if currently unplaced (protect actual placements done via workflow)
                    if existing.get_norm_placement_status() != 'PLACED':
                        existing.placement_status = placement_status
                    updated_count += 1
                else:
                    max_sno += 1
                    student = Student(
                        s_no=max_sno,
                        reg_no=reg_no,
                        name=name,
                        department=dept,
                        gender=gender,
                        hosteller_status=hosteller,
                        sslc_percentage=sslc_pct,
                        hsc_percentage=hsc_pct,
                        ug_percentage=ug_pct,
                        pg_percentage=pg_pct,
                        diploma_percentage=diploma_pct,
                        current_arrears=curr_arr,
                        history_arrears=hist_arr,
                        graduation_year=grad_year,
                        github_id=github_id,
                        linkedin_id=linkedin_id,
                        resume_link=resume_link,
                        self_intro_link=self_intro_link,
                        photo_link=photo_link,
                        portfolio_link=portfolio_link,
                        email=email,
                        phone=phone,
                        placement_status=placement_status,
                        placed_company='N/A',
                        salary_package='N/A',
                        remarks=''
                    )
                    db.session.add(student)
                    added_count += 1
            except Exception as row_err:
                row_errors.append({'row': int(row_idx) + 2, 'reg_no': reg_no, 'error': str(row_err)})
                continue

        db.session.commit()

        # Post-import stats
        total_students = Student.query.count()
        placed_count = Student.query.filter(
            (Student.placement_status == 'PLACED') | (Student.placement_status == 'YES')
        ).count()
        yet_to_be_placed_count = total_students - placed_count

        result = {
            'message': f'{added_count} students imported successfully.' if added_count > 0 else f'{updated_count} students updated successfully.',
            'added': added_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'errors': row_errors,
            'stats': {
                'total_students': total_students,
                'placed': placed_count,
                'yet_to_be_placed': yet_to_be_placed_count,
            }
        }
        return jsonify(result), 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] upload_excel failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to process spreadsheet file', 'details': str(e)}), 500


@students_bp.route('/template', methods=['GET'])
def download_template():
    """
    Download a blank Student Excel template matching the 100_Students_List.xlsx first sheet structure.
    Columns exactly match what upload_excel expects.
    """
    headers = [
        'Roll No',
        'Name',
        'Department',
        'Gender',
        'Student Type',
        'SSLC %',
        'HSC %',
        'UG %',
        'PG %',
        'GitHub ID',
        'Resume Link',
        'LinkedIn ID',
        'Graduation Date',
        'Portfolio',
        'Personal Email ID',
        'College Email ID',
        'Mobile No',
        'Student Photo',
        'Placement Status',
    ]
    df = pd.DataFrame(columns=headers)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Students Directory')
        # Style: freeze header row and auto-width columns
        ws = writer.sheets['Students Directory']
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color='1E4F91', end_color='1E4F91', fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 32
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(14, min(max_len + 4, 36))
        ws.freeze_panes = 'A2'

    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Student_Details_Template.xlsx'
    )

