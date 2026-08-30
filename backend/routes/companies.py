import os
import re
import io
import uuid
import secrets
from datetime import datetime
import traceback
import pandas as pd
from werkzeug.utils import secure_filename
from flask import Blueprint, request, jsonify, send_file, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt
from app.extensions import db
from app.models import Company, Student, User, CompanyRegistration

companies_bp = Blueprint('companies', __name__)

COMPANY_STATUSES = ['Cold', 'Warm', 'Hot', 'Drive Completed']
APPROVAL_STATUSES = ['PENDING', 'APPROVED', 'REJECTED']

ALLOWED_JD_EXTENSIONS = {'pdf', 'doc', 'docx'}
MAX_JD_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

def get_current_user_role():
    """Helper to get authenticated user role from JWT claims or Database"""
    try:
        claims = get_jwt()
        if claims and 'role' in claims:
            return claims['role']
    except Exception:
        pass

    try:
        user_id = get_jwt_identity()
        if user_id:
            user = db.session.get(User, int(user_id))
            if user:
                return user.role
    except Exception:
        pass

    return None

def get_current_user_display():
    """Helper to get authenticated user display string (Name (ROLE))"""
    try:
        user_id = get_jwt_identity()
        if user_id:
            user = db.session.get(User, int(user_id))
            if user:
                return f"{user.name} ({user.role})"
    except Exception:
        pass
    return "Faculty Member"

def allowed_jd_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_JD_EXTENSIONS

def get_jd_upload_dir():
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'uploads', 'company_jds'))
    os.makedirs(base_dir, exist_ok=True)
    return base_dir

def handle_jd_file_upload(file):
    """Validates, securely stores JD file, and returns (relative_path, original_filename)"""
    if not file or not file.filename:
        return None, None, None

    filename = file.filename
    if not allowed_jd_file(filename):
        return None, None, 'Please upload a valid JD document (PDF, DOC, or DOCX).'

    # Check file size
    file.seek(0, os.SEEK_END)
    file_length = file.tell()
    file.seek(0)

    if file_length > MAX_JD_FILE_SIZE:
        return None, None, 'JD file is too large. Maximum allowed size is 10 MB.'

    # Safe unique filename
    orig_secure = secure_filename(filename) or 'job_description.pdf'
    ext = orig_secure.rsplit('.', 1)[1].lower() if '.' in orig_secure else 'pdf'
    unique_name = f"jd_{uuid.uuid4().hex[:12]}_{orig_secure}"
    
    upload_dir = get_jd_upload_dir()
    full_path = os.path.join(upload_dir, unique_name)
    file.save(full_path)

    # Relative path stored in DB
    rel_path = os.path.join('uploads', 'company_jds', unique_name).replace('\\', '/')
    return rel_path, filename, None

def validate_company_payload(data, is_update=False):
    errors = {}
    
    # 1. Company Name
    name = str(data.get('company_name') or data.get('name') or '').strip()
    if not is_update or ('company_name' in data or 'name' in data):
        if not name:
            errors['company_name'] = 'Company name is required.'

    # 2. Location (City / Region text)
    location = str(data.get('location') or '').strip()
    if not is_update or 'location' in data:
        if not location:
            errors['location'] = 'Location is required.'

    # 3. Google Maps Location Link (Required URL)
    maps_link = str(data.get('google_maps_link') or data.get('company_address') or data.get('maps_link') or '').strip()
    if not is_update or ('google_maps_link' in data or 'company_address' in data or 'maps_link' in data):
        if not maps_link:
            errors['google_maps_link'] = 'Google Maps Location Link is required.'
        else:
            url_pattern = r'^(https?:\/\/)?([\da-z\.-]+)\.([a-z\.]{2,6})([\/\w \.-]*)*\/?(\?.*)?$'
            is_valid_url = bool(re.match(url_pattern, maps_link, re.IGNORECASE)) or maps_link.startswith('http://') or maps_link.startswith('https://') or 'maps.google' in maps_link or 'goo.gl' in maps_link or 'maps.app.goo.gl' in maps_link
            if not is_valid_url:
                errors['google_maps_link'] = 'Please enter a valid Google Maps location URL (e.g. https://maps.google.com/...).'

    # 4. Relationship Status
    status = str(data.get('status') or 'Cold').strip()
    if not is_update or 'status' in data:
        if status not in COMPANY_STATUSES:
            errors['status'] = f'Status must be one of: {", ".join(COMPANY_STATUSES)}.'

    # 5. Website validation if provided
    website = str(data.get('website') or '').strip()
    if website:
        if not (website.startswith('http://') or website.startswith('https://') or '.' in website):
            errors['website'] = 'Please enter a valid website URL.'

    # 6. Contact Person Email validation if provided
    email = str(data.get('contact_person_email') or data.get('contact_person_mail') or data.get('contact_email') or '').strip()
    if email:
        email_pattern = r'^[^@\s]+@[^@\s]+\.[^@\s]+$'
        if not re.match(email_pattern, email):
            errors['contact_person_email'] = 'Please enter a valid contact email address.'

    # 7. Contact Person Phone validation if provided
    phone = str(data.get('contact_person_number') or data.get('contact_phone') or '').strip()
    if phone:
        digits_only = re.sub(r'\D', '', phone)
        if len(digits_only) < 7 or len(digits_only) > 15:
            errors['contact_person_number'] = 'Please enter a valid contact phone number (7-15 digits).'

    # 8. No. of Hirings (Required integer >= 0)
    hirings_raw = data.get('no_of_hirings') if data.get('no_of_hirings') is not None else data.get('employee_count')
    hirings_val = 0
    if not is_update or ('no_of_hirings' in data or 'employee_count' in data):
        if hirings_raw is None or str(hirings_raw).strip() == '':
            errors['no_of_hirings'] = 'Number of hirings is required.'
        else:
            try:
                hirings_val = int(hirings_raw)
                if hirings_val < 0:
                    errors['no_of_hirings'] = 'Number of hirings cannot be negative.'
            except (ValueError, TypeError):
                errors['no_of_hirings'] = 'Number of hirings must be a valid integer.'

    # 9. CTC (LPA) (Required numeric > 0)
    ctc_raw = data.get('ctc_lpa') if data.get('ctc_lpa') is not None else data.get('ctc')
    if not is_update or ('ctc_lpa' in data or 'ctc' in data):
        if ctc_raw is None or str(ctc_raw).strip() == '':
            errors['ctc_lpa'] = 'CTC (LPA) is required.'
        else:
            try:
                ctc_val = float(ctc_raw)
                if ctc_val <= 0:
                    errors['ctc_lpa'] = 'CTC (LPA) must be greater than 0.'
            except (ValueError, TypeError):
                errors['ctc_lpa'] = 'CTC (LPA) must be a valid number (e.g. 4.5, 6.0, 12.0).'

    return errors

@companies_bp.route('/', methods=['GET'])
def get_companies():
    try:
        status = request.args.get('status')
        approval_status = request.args.get('approval_status')
        search = request.args.get('search')

        query = Company.query
        if status:
            query = query.filter(Company.status == status)
        if approval_status:
            query = query.filter(Company.approval_status == approval_status.upper())
        if search:
            query = query.filter(
                (Company.name.ilike(f'%{search}%')) |
                (Company.location.ilike(f'%{search}%')) |
                (Company.google_maps_link.ilike(f'%{search}%')) |
                (Company.contact_person_number.ilike(f'%{search}%')) |
                (Company.contact_person_email.ilike(f'%{search}%'))
            )

        companies = query.order_by(Company.created_at.desc(), Company.id.desc()).all()

        # Batch pre-fetch placed counts for all companies in 1 query
        placed_counts = dict(
            db.session.query(
                Student.placed_company_id,
                db.func.count(Student.id)
            ).filter(
                Student.placed_company_id.isnot(None),
                (Student.placement_status == 'PLACED') | (Student.placement_status == 'Placed') | (Student.placement_status == 'YES')
            ).group_by(Student.placed_company_id).all()
        )

        # Batch pre-fetch registration counts for all companies in 1 query
        reg_counts = dict(
            db.session.query(
                CompanyRegistration.company_id,
                db.func.count(CompanyRegistration.id)
            ).filter(
                CompanyRegistration.registration_status == 'REGISTERED'
            ).group_by(CompanyRegistration.company_id).all()
        )

        return jsonify([
            c.to_dict(
                placed_count=placed_counts.get(c.id, 0),
                registered_count=reg_counts.get(c.id, 0)
            ) for c in companies
        ]), 200
    except Exception as e:
        print(f"[ERROR] get_companies failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to retrieve companies', 'details': str(e)}), 500

@companies_bp.route('/template', methods=['GET'])
def download_company_template():
    """
    Download a blank Excel template for Company Details matching the Companies_List(1).xlsx structure.
    Contains ONLY the 14 standard headers with no sample data.
    """
    try:
        headers = [
            'S.No',
            'Company Name',
            'Job Title / Role',
            'CTC (LPA)',
            'Location',
            'Opportunity Status',
            'Job Status',
            'Placed Students Count',
            'Placed Students Details',
            'Job Description Summary',
            'JD PDF Link',
            'Official Careers Link',
            'Contact Email',
            'Contact Mobile'
        ]
        df = pd.DataFrame(columns=headers)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Companies & Job Drives')
            # Style: freeze header row and auto-width columns
            ws = writer.sheets['Companies & Job Drives']
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color='1E4F91', end_color='1E4F91', fill_type='solid')
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[1].height = 32
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max(14, min(max_len + 4, 36))
            ws.freeze_panes = 'A2'

        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Company_Details_Template.xlsx'
        )
    except Exception as e:
        print(f"[ERROR] download_company_template failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to generate company template', 'details': str(e)}), 500

@companies_bp.route('/export', methods=['GET'])
@jwt_required()
def export_companies():
    """Export filtered company details to Excel (.xlsx) using actual placement records"""
    try:
        status = request.args.get('status')
        approval_status = request.args.get('approval_status')
        search = request.args.get('search')

        query = Company.query
        if status and status.strip() and status != 'ALL':
            query = query.filter(Company.status == status.strip())
        if approval_status and approval_status.strip() and approval_status != 'ALL':
            query = query.filter(Company.approval_status == approval_status.strip().upper())
        if search and search.strip():
            query = query.filter(
                (Company.name.ilike(f'%{search.strip()}%')) |
                (Company.location.ilike(f'%{search.strip()}%')) |
                (Company.google_maps_link.ilike(f'%{search.strip()}%')) |
                (Company.contact_person_number.ilike(f'%{search.strip()}%')) |
                (Company.contact_person_email.ilike(f'%{search.strip()}%'))
            )

        companies = query.order_by(Company.created_at.desc(), Company.id.desc()).all()

        if not companies:
            return jsonify({'error': 'No company data available to export.'}), 404

        # Prepare 14 columns matching the standard company template
        rows = []
        for idx, comp in enumerate(companies, start=1):
            is_completed = (comp.status == 'Drive Completed')
            real_placed = comp.get_real_placed_count()
            placed_val = real_placed if is_completed else 0

            rows.append({
                'S.No': idx,
                'Company Name': comp.name,
                'Job Title / Role': comp.job_title or comp.industry or 'Software Engineer',
                'CTC (LPA)': comp.ctc_lpa if comp.ctc_lpa is not None else '',
                'Location': comp.location or 'N/A',
                'Opportunity Status': comp.status,
                'Job Status': comp.job_status or ('Completed' if is_completed else 'Active'),
                'Placed Students Count': placed_val,
                'Placed Students Details': f"{placed_val} students placed" if placed_val > 0 else '',
                'Job Description Summary': comp.jd_summary or comp.remarks or '',
                'JD PDF Link': comp.jd_pdf_link or comp.jd_file_path or '',
                'Official Careers Link': comp.website or '',
                'Contact Email': comp.contact_person_email or comp.contact_email or '',
                'Contact Mobile': comp.contact_person_number or comp.contact_phone or ''
            })

        df = pd.DataFrame(rows)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Companies & Job Drives')
        output.seek(0)

        filename = f"Company_Details_{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"[ERROR] export_companies failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to export company details', 'details': str(e)}), 500

@companies_bp.route('/import', methods=['POST'])
@jwt_required()
def import_companies():
    """
    Import company details from Excel (.xlsx/.xls) file.
    Imports ONLY the first sheet (e.g. 'Companies & Job Drives').
    Sheets 2 & 3 (Students Directory, Placements & Drive History) are ALWAYS ignored.
    Does NOT create fake placement records from Placed Students Count.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded. Please provide an Excel (.xlsx) file.'}), 400

    file = request.files['file']
    filename = file.filename.lower()

    if not (filename.endswith('.xlsx') or filename.endswith('.xls') or filename.endswith('.csv')):
        return jsonify({'error': 'Invalid file format. Please upload an Excel (.xlsx) file.'}), 400

    COLUMN_ALIASES = {
        'name':            ['Company Name', 'Company', 'Name', 'company_name', 'Organization', 'Employer'],
        'job_title':       ['Job Title / Role', 'Job Title', 'Role', 'Job Role', 'Designation', 'Position', 'job_title', 'job_role'],
        'ctc_lpa':         ['CTC (LPA)', 'CTC', 'CTC LPA', 'Package', 'ctc_lpa', 'package_offered', 'Salary'],
        'location':        ['Location', 'City', 'Job Location', 'location', 'Work Location'],
        'status':          ['Opportunity Status', 'Relationship Status', 'Status', 'Opportunity', 'status'],
        'job_status':      ['Job Status', 'Drive Status', 'job_status', 'Posting Status'],
        'jd_summary':      ['Job Description Summary', 'JD Summary', 'Description', 'Job Summary', 'jd_summary', 'Remarks', 'remarks'],
        'jd_pdf_link':     ['JD PDF Link', 'JD Link', 'JD File', 'jd_pdf_link', 'jd_file_path', 'PDF Link'],
        'website':         ['Official Careers Link', 'Careers Link', 'Website', 'Website URL', 'website', 'careers_url', 'Portal Link'],
        'contact_email':   ['Contact Email', 'Contact Person Mail', 'Contact Person Email', 'Email', 'contact_person_email', 'contact_email'],
        'contact_mobile':  ['Contact Mobile', 'Contact Person Number', 'Contact Number', 'Contact Phone', 'Phone', 'Mobile', 'contact_person_number', 'contact_phone'],
        'maps_link':       ['Google Maps Location Link', 'Google Maps Link', 'Maps Link', 'google_maps_link', 'company_address', 'Maps URL'],
    }

    def normalize_col(col):
        return str(col).strip().lower()

    def find_col(headers_lower, aliases):
        for alias in aliases:
            a_lower = normalize_col(alias)
            if a_lower in headers_lower:
                return headers_lower[a_lower]
        return None

    def cell_val(row, col_idx, default=None):
        if col_idx is None:
            return default
        try:
            val = row.iloc[col_idx]
            if pd.isna(val):
                return default
            return val
        except Exception:
            return default

    def str_val(row, col_idx, default=''):
        v = cell_val(row, col_idx)
        if v is None:
            return default
        return str(v).strip()

    try:
        file_bytes = file.read()
        import io as _io

        if filename.endswith('.csv'):
            df_raw = pd.read_csv(_io.BytesIO(file_bytes), header=None, dtype=str)
        else:
            # Read ONLY the first sheet; ignore all other sheets (e.g. Students Directory, Placements & Drive History)
            xl = pd.ExcelFile(_io.BytesIO(file_bytes))
            first_sheet = xl.sheet_names[0]
            df_raw = xl.parse(first_sheet, header=None, dtype=str)

        if df_raw.empty:
            return jsonify({'error': 'The uploaded Excel file is empty.'}), 400

        # --- Locate header row within first 10 rows ---
        HEADER_CLUES = {'company name', 'company', 'job title', 'job title / role', 'ctc (lpa)', 'location', 'opportunity status'}
        header_row_idx = None
        for i in range(min(10, len(df_raw))):
            row_vals = [normalize_col(c) for c in df_raw.iloc[i].fillna('').tolist()]
            if any(clue in row_vals for clue in HEADER_CLUES):
                header_row_idx = i
                break

        if header_row_idx is None:
            # Fallback to row 0 if no clue matched
            header_row_idx = 0

        headers_raw = df_raw.iloc[header_row_idx].fillna('').tolist()
        headers_lower = {normalize_col(h): idx for idx, h in enumerate(headers_raw)}
        df_data = df_raw.iloc[header_row_idx + 1:].reset_index(drop=True)

        col_map = {}
        for field, aliases in COLUMN_ALIASES.items():
            col_map[field] = find_col(headers_lower, aliases)

        if col_map['name'] is None:
            return jsonify({
                'error': 'Required column "Company Name" was not found in the uploaded file.',
                'details': f'Detected headers: {list(headers_raw)}'
            }), 400

        added_count = 0
        updated_count = 0
        user_display = get_current_user_display()

        for idx, row in df_data.iterrows():
            # Skip completely blank rows
            if row.dropna().empty or all(str(v).strip() == '' for v in row.tolist()):
                continue

            name = str_val(row, col_map['name'])
            if not name:
                continue

            job_title = str_val(row, col_map['job_title'], 'Software Engineer') or 'Software Engineer'
            location = str_val(row, col_map['location'], 'Chennai') or 'Chennai'
            
            # CTC (LPA)
            ctc_raw = str_val(row, col_map['ctc_lpa'])
            ctc_val = None
            if ctc_raw:
                try:
                    ctc_clean = str(ctc_raw).upper().replace('LPA', '').replace('L', '').strip()
                    ctc_val = float(ctc_clean)
                    if ctc_val <= 0:
                        ctc_val = 6.0
                except (ValueError, TypeError):
                    ctc_val = 6.0
            else:
                ctc_val = 6.0

            # Relationship status
            status_raw = str_val(row, col_map['status'], 'Cold')
            matched_status = 'Cold'
            for valid_s in COMPANY_STATUSES:
                if valid_s.lower() == status_raw.lower():
                    matched_status = valid_s
                    break

            job_status = str_val(row, col_map['job_status'], 'Active') or 'Active'
            jd_summary = str_val(row, col_map['jd_summary'])
            jd_pdf_link = str_val(row, col_map['jd_pdf_link'])
            website = str_val(row, col_map['website'])
            contact_email = str_val(row, col_map['contact_email'])
            contact_mobile = str_val(row, col_map['contact_mobile'])

            # Google maps link fallback
            maps_link = str_val(row, col_map['maps_link'])
            if not maps_link:
                maps_link = f"https://maps.google.com/?q={name.replace(' ', '+')}+{location.replace(' ', '+')}"

            # Always start with PENDING approval for imported companies
            approval_status = 'PENDING'

            # Duplicate check: by company name (case-insensitive)
            existing = Company.query.filter(Company.name.ilike(name.strip())).first()

            if existing:
                existing.job_title = job_title
                existing.industry = job_title
                existing.location = location
                existing.ctc_lpa = ctc_val
                existing.package_offered = f"{ctc_val} LPA"
                existing.status = matched_status
                existing.job_status = job_status
                if jd_summary:
                    existing.jd_summary = jd_summary
                    existing.remarks = jd_summary
                if jd_pdf_link:
                    existing.jd_pdf_link = jd_pdf_link
                if website:
                    existing.website = website
                if contact_email:
                    existing.contact_person_email = contact_email
                if contact_mobile:
                    existing.contact_person_number = contact_mobile
                if maps_link and not existing.google_maps_link:
                    existing.google_maps_link = maps_link
                # Do NOT create fake placement counts from Excel
                existing.placed_students = existing.get_real_placed_count()
                updated_count += 1
            else:
                new_comp = Company(
                    name=name,
                    job_title=job_title,
                    industry=job_title,
                    location=location,
                    ctc_lpa=ctc_val,
                    package_offered=f"{ctc_val} LPA",
                    no_of_hirings=5, # default reasonable hiring capacity
                    employee_count=5,
                    status=matched_status,
                    job_status=job_status,
                    approval_status=approval_status,
                    jd_summary=jd_summary,
                    remarks=jd_summary,
                    jd_pdf_link=jd_pdf_link,
                    website=website,
                    contact_person_email=contact_email,
                    contact_person_number=contact_mobile,
                    google_maps_link=maps_link,
                    company_address=maps_link,
                    placed_students=0, # strictly 0 until real student placement occurs
                    created_by_user=user_display
                )
                db.session.add(new_comp)
                added_count += 1

        db.session.commit()

        return jsonify({
            'message': f"Successfully processed. Added {added_count} new companies, updated {updated_count} existing companies (submitted for approval).",
            'added': added_count,
            'updated': updated_count,
            'count': added_count + updated_count
        }), 201

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] import_companies failed: {traceback.format_exc()}")
        return jsonify({'error': f"Failed to import Excel file: {str(e)}"}), 500

@companies_bp.route('/<int:company_id>', methods=['GET'])
def get_company(company_id):
    try:
        company = db.session.get(Company, company_id)
        if not company:
            return jsonify({'error': 'Company not found'}), 404
        return jsonify(company.to_dict()), 200
    except Exception as e:
        print(f"[ERROR] get_company failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to retrieve company', 'details': str(e)}), 500

@companies_bp.route('/<int:company_id>/jd', methods=['GET'])
def get_company_jd(company_id):
    """Secure endpoint to view or download company Job Description (JD)"""
    try:
        company = db.session.get(Company, company_id)
        if not company:
            return jsonify({'error': 'Company not found'}), 404

        if not company.jd_file_path:
            return jsonify({'error': 'Job Description (JD) not available for this company.'}), 404

        # Resolve path safely
        backend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        file_path = os.path.abspath(os.path.join(backend_dir, company.jd_file_path))

        if not os.path.exists(file_path):
            return jsonify({'error': 'Job Description file not found on server.'}), 404

        download = request.args.get('download', '').lower() in ('1', 'true', 'yes')
        filename = company.jd_file_name or os.path.basename(file_path)

        # Determine mimetype
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        mimetype_map = {
            'pdf': 'application/pdf',
            'doc': 'application/msword',
            'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        }
        mimetype = mimetype_map.get(ext, 'application/octet-stream')

        return send_file(
            file_path,
            mimetype=mimetype,
            as_attachment=download,
            download_name=filename
        )

    except Exception as e:
        print(f"[ERROR] get_company_jd failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to retrieve Job Description', 'details': str(e)}), 500

@companies_bp.route('/', methods=['POST'])
def add_company():
    """Create company with optional JD document upload"""
    try:
        # Support both JSON payload and multipart/form-data
        if request.is_json:
            data = request.get_json() or {}
        else:
            data = request.form.to_dict()

        # Validation
        errors = validate_company_payload(data, is_update=False)
        if errors:
            first_err = next(iter(errors.values()))
            return jsonify({'error': first_err, 'validation_errors': errors}), 400

        # Handle optional JD file upload
        jd_file_path = None
        jd_file_name = None
        if 'jd_file' in request.files:
            file = request.files['jd_file']
            if file and file.filename:
                rel_path, orig_name, err = handle_jd_file_upload(file)
                if err:
                    return jsonify({'error': err}), 400
                jd_file_path = rel_path
                jd_file_name = orig_name

        name = str(data.get('company_name') or data.get('name') or '').strip()
        location = str(data.get('location') or '').strip()
        website = str(data.get('website') or '').strip()
        phone = str(data.get('contact_person_number') or data.get('contact_phone') or '').strip()
        email = str(data.get('contact_person_email') or data.get('contact_person_mail') or data.get('contact_email') or '').strip()
        
        hirings_raw = data.get('no_of_hirings') if data.get('no_of_hirings') is not None else data.get('employee_count')
        no_of_hirings = int(hirings_raw) if (hirings_raw is not None and str(hirings_raw).strip() != '') else 0
        
        ctc_raw = data.get('ctc_lpa') if data.get('ctc_lpa') is not None else data.get('ctc')
        ctc_lpa = float(ctc_raw) if (ctc_raw is not None and str(ctc_raw).strip() != '') else None
        
        maps_link = str(data.get('google_maps_link') or data.get('company_address') or data.get('maps_link') or '').strip()
        status = str(data.get('status') or 'Cold').strip()
        
        # Initialized to 0; derived strictly from real student records
        placed_students = 0
        approval_status = 'PENDING'
        
        # Metadata
        industry = str(data.get('industry') or 'Technology').strip()
        contact_person = str(data.get('contact_person') or '').strip()
        package_offered = f"{ctc_lpa} LPA" if ctc_lpa is not None else str(data.get('package_offered') or '').strip()
        drive_date = str(data.get('drive_date') or '').strip()
        remarks = str(data.get('remarks') or '').strip()
        faculty_in_charge = str(data.get('faculty_in_charge') or 'Unassigned').strip()
        created_by_user = str(data.get('created_by_user') or 'Faculty User').strip()

        company = Company(
            name=name,
            location=location,
            website=website,
            contact_person_number=phone,
            contact_person_email=email,
            no_of_hirings=no_of_hirings,
            employee_count=no_of_hirings,
            ctc_lpa=ctc_lpa,
            placed_students=placed_students,
            google_maps_link=maps_link,
            company_address=maps_link,
            jd_file_path=jd_file_path,
            jd_file_name=jd_file_name,
            status=status,
            approval_status=approval_status,
            industry=industry,
            contact_person=contact_person,
            package_offered=package_offered,
            drive_date=drive_date,
            remarks=remarks,
            faculty_in_charge=faculty_in_charge,
            created_by_user=created_by_user
        )

        db.session.add(company)
        db.session.commit()
        return jsonify({
            'message': 'Company submitted for admin approval.',
            'company': company.to_dict()
        }), 201

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] add_company failed: {traceback.format_exc()}")
        return jsonify({'error': 'Unable to save company. Please try again.', 'details': str(e)}), 500

@companies_bp.route('/<int:company_id>', methods=['PUT'])
def update_company(company_id):
    """Update company with support for replacing or removing JD document"""
    try:
        company = db.session.get(Company, company_id)
        if not company:
            return jsonify({'error': 'Company not found'}), 404

        if request.is_json:
            data = request.get_json() or {}
        else:
            data = request.form.to_dict()
        
        # Merge existing fields with incoming data for cross-field validation
        merged = company.to_dict()
        merged.update(data)
        
        # Validation
        errors = validate_company_payload(merged, is_update=True)
        if errors:
            first_err = next(iter(errors.values()))
            return jsonify({'error': first_err, 'validation_errors': errors}), 400

        # Handle removing JD if requested
        remove_jd = str(data.get('remove_jd') or '').lower() in ('true', '1', 'yes')
        if remove_jd and company.jd_file_path:
            try:
                backend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
                old_file = os.path.join(backend_dir, company.jd_file_path)
                if os.path.exists(old_file):
                    os.remove(old_file)
            except Exception as e:
                print(f"[WARN] Failed to delete old JD file: {e}")
            company.jd_file_path = None
            company.jd_file_name = None

        # Handle new/replacement JD file upload
        if 'jd_file' in request.files:
            file = request.files['jd_file']
            if file and file.filename:
                rel_path, orig_name, err = handle_jd_file_upload(file)
                if err:
                    return jsonify({'error': err}), 400
                
                # Delete previous JD file if replaced
                if company.jd_file_path:
                    try:
                        backend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
                        old_file = os.path.join(backend_dir, company.jd_file_path)
                        if os.path.exists(old_file):
                            os.remove(old_file)
                    except Exception as e:
                        print(f"[WARN] Failed to delete replaced JD file: {e}")

                company.jd_file_path = rel_path
                company.jd_file_name = orig_name

        if 'company_name' in data or 'name' in data:
            company.name = str(data.get('company_name') or data.get('name')).strip()
        if 'location' in data:
            company.location = str(data['location']).strip()
        if 'website' in data:
            company.website = str(data['website']).strip()
        if 'contact_person_number' in data or 'contact_phone' in data:
            company.contact_person_number = str(data.get('contact_person_number') or data.get('contact_phone')).strip()
        if 'contact_person_email' in data or 'contact_person_mail' in data or 'contact_email' in data:
            company.contact_person_email = str(data.get('contact_person_email') or data.get('contact_person_mail') or data.get('contact_email')).strip()
        if 'no_of_hirings' in data or 'employee_count' in data:
            hirings_raw = data.get('no_of_hirings') if data.get('no_of_hirings') is not None else data.get('employee_count')
            company.no_of_hirings = int(hirings_raw) if (hirings_raw is not None and str(hirings_raw).strip() != '') else 0
            company.employee_count = company.no_of_hirings
        if 'ctc_lpa' in data or 'ctc' in data:
            ctc_raw = data.get('ctc_lpa') if data.get('ctc_lpa') is not None else data.get('ctc')
            company.ctc_lpa = float(ctc_raw) if (ctc_raw is not None and str(ctc_raw).strip() != '') else None
            company.package_offered = f"{company.ctc_lpa} LPA" if company.ctc_lpa is not None else 'N/A'
        if 'google_maps_link' in data or 'company_address' in data or 'maps_link' in data:
            maps_link = str(data.get('google_maps_link') or data.get('company_address') or data.get('maps_link')).strip()
            company.google_maps_link = maps_link
            company.company_address = maps_link
        if 'status' in data:
            company.status = str(data['status']).strip()
        
        # Placed students is strictly derived from actual student placement records
        company.placed_students = company.get_real_placed_count()
            
        if 'industry' in data:
            company.industry = str(data['industry']).strip()
        if 'contact_person' in data:
            company.contact_person = str(data['contact_person']).strip()
        if 'package_offered' in data and 'ctc_lpa' not in data:
            company.package_offered = str(data['package_offered']).strip()
        if 'drive_date' in data:
            company.drive_date = str(data['drive_date']).strip()
        if 'remarks' in data:
            company.remarks = str(data['remarks']).strip()
        if 'faculty_in_charge' in data:
            company.faculty_in_charge = str(data['faculty_in_charge']).strip()

        db.session.commit()
        return jsonify({'message': 'Company updated successfully', 'company': company.to_dict()}), 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] update_company failed: {traceback.format_exc()}")
        return jsonify({'error': 'Unable to update company. Please try again.', 'details': str(e)}), 500

@companies_bp.route('/<int:company_id>/approve', methods=['PATCH', 'PUT'])
@jwt_required()
def approve_company(company_id):
    """Admin exclusive endpoint to approve a company"""
    role = get_current_user_role()
    if role != 'ADMIN':
        return jsonify({'error': 'Forbidden: Only Admin has authority to approve companies.'}), 403

    try:
        company = db.session.get(Company, company_id)
        if not company:
            return jsonify({'error': 'Company not found'}), 404

        company.approval_status = 'APPROVED'
        if not company.registration_token:
            company.registration_token = uuid.uuid4().hex[:16]
        company.registration_link_status = 'ACTIVE'

        db.session.commit()
        return jsonify({
            'message': f'Company "{company.name}" has been approved successfully.',
            'company': company.to_dict(),
            'registration_token': company.registration_token
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] approve_company failed: {traceback.format_exc()}")
        return jsonify({'error': 'Unable to approve company. Please try again.', 'details': str(e)}), 500

@companies_bp.route('/<int:company_id>/reject', methods=['PATCH', 'PUT'])
@jwt_required()
def reject_company(company_id):
    """Admin exclusive endpoint to reject a company"""
    role = get_current_user_role()
    if role != 'ADMIN':
        return jsonify({'error': 'Forbidden: Only Admin has authority to reject companies.'}), 403

    try:
        company = db.session.get(Company, company_id)
        if not company:
            return jsonify({'error': 'Company not found'}), 404

        company.approval_status = 'REJECTED'
        company.registration_link_status = 'INACTIVE'

        db.session.commit()
        return jsonify({
            'message': f'Company "{company.name}" has been rejected.',
            'company': company.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] reject_company failed: {traceback.format_exc()}")
        return jsonify({'error': 'Unable to reject company. Please try again.', 'details': str(e)}), 500

@companies_bp.route('/<int:company_id>/registrations', methods=['GET'])
@jwt_required()
def get_company_registrations(company_id):
    """Authenticated endpoint to view all students registered for a company drive"""
    try:
        company = db.session.get(Company, company_id)
        if not company:
            return jsonify({'error': 'Company not found'}), 404

        search = request.args.get('search', '').strip().lower()
        dept = request.args.get('department', '').strip().upper()
        gender = request.args.get('gender', '').strip()
        student_type = request.args.get('student_type', '').strip()
        reg_status = request.args.get('registration_status', '').strip()
        placement_status = request.args.get('placement_status', '').strip().upper()

        regs = CompanyRegistration.query.filter_by(company_id=company.id).order_by(CompanyRegistration.registered_at.desc()).all()
        result = []

        for idx, r in enumerate(regs, start=1):
            s = r.student
            if not s:
                continue

            # Filtering
            if search:
                match_search = (
                    search in (s.reg_no or '').lower() or
                    search in (s.name or '').lower() or
                    search in (s.department or s.dept or '').lower()
                )
                if not match_search:
                    continue

            if dept and (s.department or s.dept or '').upper() != dept:
                continue

            if gender and s.gender != gender:
                continue

            if student_type and (s.hosteller_status or s.hosteller_day_scholar or '') != student_type:
                continue

            if reg_status and r.registration_status != reg_status:
                continue

            if placement_status:
                norm_p = s.get_norm_placement_status()
                if placement_status == 'PLACED' and norm_p != 'PLACED':
                    continue
                if placement_status in ('YET_TO_BE_PLACED', 'UNPLACED') and norm_p == 'PLACED':
                    continue

            r_dict = r.to_dict()
            r_dict['s_no'] = idx
            result.append(r_dict)

        return jsonify({
            'company': company.to_dict(),
            'registrations': result,
            'total_registered': company.get_registered_students_count()
        }), 200

    except Exception as e:
        print(f"[ERROR] get_company_registrations failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to retrieve company registrations', 'details': str(e)}), 500

@companies_bp.route('/<int:company_id>/registration-link', methods=['GET', 'POST'])
@jwt_required()
def get_or_create_registration_link(company_id):
    """
    Returns registration link for an APPROVED company.
    If token doesn't exist, generates a unique secure token and persists in Supabase.
    """
    try:
        company = db.session.get(Company, company_id)
        if not company:
            return jsonify({'error': 'Company not found'}), 404

        if company.approval_status != 'APPROVED':
            return jsonify({
                'error': 'Registration unavailable until company approval.',
                'approval_status': company.approval_status,
                'is_active': False
            }), 403

        if not company.registration_token:
            company.registration_token = secrets.token_hex(8)
            company.registration_link_status = 'ACTIVE'
            db.session.commit()
        elif company.registration_link_status != 'ACTIVE':
            company.registration_link_status = 'ACTIVE'
            db.session.commit()

        return jsonify({
            'company_id': company.id,
            'company_name': company.name,
            'registration_token': company.registration_token,
            'registration_link_status': company.registration_link_status,
            'registered_students_count': company.get_registered_students_count(),
            'is_active': True
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] registration-link failed: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to fetch registration link', 'details': str(e)}), 500

@companies_bp.route('/<int:company_id>', methods=['DELETE'])
def delete_company(company_id):
    try:
        company = db.session.get(Company, company_id)
        if not company:
            return jsonify({'error': 'Company not found'}), 404

        # Clean up JD file if present
        if company.jd_file_path:
            try:
                backend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
                old_file = os.path.join(backend_dir, company.jd_file_path)
                if os.path.exists(old_file):
                    os.remove(old_file)
            except Exception as e:
                print(f"[WARN] Failed to delete JD file on company deletion: {e}")
            
        db.session.delete(company)
        db.session.commit()
        return jsonify({'message': 'Company deleted successfully'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] delete_company failed: {traceback.format_exc()}")
        return jsonify({'error': 'Unable to delete company. Please try again.', 'details': str(e)}), 500
